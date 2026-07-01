"""Excel export service — Farmer Statement and Collection Summary."""
from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def _t(k,**kw):
    from translations import t; return t(k,**kw)
def _bs(d,lang):
    from utils.bs_calendar import db_date_to_bs_str
    return db_date_to_bs_str(d,lang=lang) if d else _t("all_time")
def _org(lang):
    from database.database import get_setting
    if lang=="NE":
        return get_setting("organization_name_nepali","") or \
               get_setting("organization_name_english","Santosh Dairy Cooperative")
    return get_setting("organization_name_english","Santosh Dairy Cooperative")

HF=PatternFill(start_color="1A2B4C",end_color="1A2B4C",fill_type="solid")
HFT=Font(bold=True,color="FFFFFF",size=10)
TF=Font(bold=True,size=14,color="1A2B4C")
SF=Font(size=10,color="6B7280")
BF=Font(bold=True,size=10)
NM="#,##0.00"

def _hr(ws,row,hdrs):
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=row,column=col,value=h)
        c.font=HFT; c.fill=HF; c.alignment=Alignment(horizontal="center")

def generate_farmer_statement_excel(statement, output_path, lang="NE"):
    wb=Workbook(); ws=wb.active
    ws.title=_t("farmer_statement")[:31]
    ws["A1"]=_org(lang); ws["A1"].font=TF; ws.merge_cells("A1:D1")
    ws["A2"]=_t("farmer_statement"); ws["A2"].font=SF; ws.merge_cells("A2:D2")
    ws["A4"]=f"{_t('col_farmer')}:"; ws["B4"]=f"{statement.farmer_code} — {statement.farmer_name}"
    next_row = 5
    if statement.bank_account:
        ws[f"A{next_row}"]=f"{_t('bank_account_label')}:"; ws[f"B{next_row}"]=statement.bank_account
        next_row += 1
    ws[f"A{next_row}"]=f"{_t('date_range')}:"; ws[f"B{next_row}"]=f"{_bs(statement.date_from,lang)} — {_bs(statement.date_to,lang)}"
    row=next_row+2
    for lbl,val in [(_t("opening_balance"),statement.opening_balance),
                    (_t("grand_total"),statement.total_transaction_amount),
                    (_t("total_paid"),statement.total_paid),
                    (_t("closing_balance"),statement.closing_balance)]:
        ws.cell(row=row,column=1,value=lbl)
        c=ws.cell(row=row,column=2,value=val); c.number_format=NM; row+=1
    row+=1
    if statement.transactions:
        ws.cell(row=row,column=1,value=_t("milk_collection")).font=Font(bold=True,size=11); row+=1
        _hr(ws,row,[_t("col_date"),_t("col_quantity"),_t("rate"),_t("col_amount")]); row+=1
        for ln in statement.transactions:
            ws.cell(row=row,column=1,value=_bs(ln.transaction_date,lang))
            ws.cell(row=row,column=2,value=ln.quantity).number_format=NM
            ws.cell(row=row,column=3,value=ln.rate).number_format=NM
            ws.cell(row=row,column=4,value=ln.amount).number_format=NM; row+=1
        row+=1
    if statement.payments:
        ws.cell(row=row,column=1,value=_t("payment_entry")).font=Font(bold=True,size=11); row+=1
        _hr(ws,row,[_t("col_date"),_t("receipt_number"),_t("amount_paid")]); row+=1
        for ln in statement.payments:
            ws.cell(row=row,column=1,value=_bs(ln.payment_date,lang))
            ws.cell(row=row,column=2,value=ln.receipt_number or "—")
            ws.cell(row=row,column=3,value=ln.amount_paid).number_format=NM; row+=1
    for col,w in zip("ABCD",[22,26,18,18]):
        ws.column_dimensions[col].width=w
    wb.save(output_path); return output_path

def generate_collection_summary_excel(summary, output_path, lang="NE"):
    wb=Workbook(); ws=wb.active
    ws.title=_t("monthly_summary")[:31]
    ws["A1"]=_org(lang); ws["A1"].font=TF; ws.merge_cells("A1:F1")
    ws["A2"]=_t("monthly_summary"); ws["A2"].font=SF; ws.merge_cells("A2:F2")
    ws["A4"]=f"{_t('date_range')}:"; ws["B4"]=f"{_bs(summary.date_from,lang)} — {_bs(summary.date_to,lang)}"
    row=6
    _hr(ws,row,[_t("col_code"),_t("col_name"),_t("col_quantity"),
                _t("col_amount"),_t("total_paid"),_t("col_balance")]); row+=1
    for r in summary.rows:
        ws.cell(row=row,column=1,value=r.farmer_code)
        ws.cell(row=row,column=2,value=r.farmer_name)
        ws.cell(row=row,column=3,value=r.total_quantity).number_format=NM
        ws.cell(row=row,column=4,value=r.total_amount).number_format=NM
        ws.cell(row=row,column=5,value=r.total_paid).number_format=NM
        ws.cell(row=row,column=6,value=r.balance).number_format=NM; row+=1
    for col,val in [(2,_t("grand_total")),(3,summary.grand_total_quantity),
                    (4,summary.grand_total_amount),(5,summary.grand_total_paid),
                    (6,summary.grand_total_balance)]:
        c=ws.cell(row=row,column=col,value=val); c.font=BF
        if col>2: c.number_format=NM
    for i,w in enumerate([14,26,16,16,16,16],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    wb.save(output_path); return output_path
